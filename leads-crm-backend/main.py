from boto3 import client
from db.client import supabase
from typing import Optional
from db.queries import (
    get_all_leads,
    get_leads_page,
    create_lead as db_create_lead,
    create_activity,
    update_lead as db_update_lead,
    get_lead_status,
    delete_lead as db_delete_lead,
    get_profile_role,
    get_profile,
    get_lead_activities as db_get_lead_activities,
)
from db.queries import upload_file_to_storage, create_attachment_record, get_lead_attachments, get_attachment, get_signed_attachment_url, delete_file_from_storage, delete_attachment_record
from models import LeadCreate, LeadUpdate, NoteCreate
from auth.permissions import require_admin
from auth.dependencies import get_current_user
from fastapi import FastAPI, Header, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import csv
import io
import json
from fastapi import UploadFile, File
from fastapi import Request
from db.client import get_client_for_user



from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi.responses import Response
from openpyxl import load_workbook
import os
import re

app = FastAPI()

from fastapi.responses import JSONResponse
from fastapi.exceptions import HTTPException as StarletteHTTPException

@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request: Request, exc: StarletteHTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail},
        headers={'Access-Control-Allow-Origin': '*'}
    )

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    traceback.print_exc()
    return JSONResponse(status_code=500, content={'detail': str(exc)}, headers={'Access-Control-Allow-Origin': '*'})
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

allowed_origins_env = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,https://crm.workfloww.ai,https://crm-git-main-workfloww.vercel.app")
allowed_origins = [origin.strip() for origin in allowed_origins_env.split(",") if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def sanitize_filename(filename: str) -> str:
    filename = filename.replace("/", "").replace("\\", "")
    filename = re.sub(r'[^a-zA-Z0-9._-]', '_', filename)
    return filename[:200]  # cap length too, just in case

@app.get("/")
def read_root():
    return {"status": "ok", "message": "Leads CRM backend is running"}

@app.get("/leads")
@limiter.limit("60/minute")
def get_leads(
    request: Request,
    page: int = 1, 
    page_size: int = 20,
    search: Optional[str] = None,
    name: Optional[str] = None,
    org: Optional[str] = None,
    title: Optional[str] = None,
    location: Optional[str] = None,
    industry: Optional[str] = None,
    function: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    user=Depends(get_current_user)
):
    sort_desc = sort_dir == "desc" if sort_dir else False
    
    db_sort_by = "due_date"
    if sort_by == "name":
        db_sort_by = "first_name"
    elif sort_by == "org":
        db_sort_by = "org"
    elif sort_by == "status":
        db_sort_by = "status"

    client = get_client_for_user(user.token)
    response = get_leads_page(client, page, page_size, search, name, org, title, location, industry, function, db_sort_by, sort_desc)
    return {"leads": response.data, "total": response.count, "page": page, "page_size": page_size}

@app.post("/leads")
@limiter.limit("60/minute")
def create_lead(request: Request, lead: LeadCreate, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    response = db_create_lead(client, lead.model_dump())
    new_lead = response.data[0]

    create_activity(client, {
        "lead_id": new_lead["id"],
        "user_id": user.id,
        "type": "created",
        "content": "Lead created",
    })

    return response.data

@app.patch("/leads/{lead_id}")
@limiter.limit("60/minute")
def update_lead(request: Request, lead_id: str, lead: LeadUpdate, user=Depends(get_current_user)):
    update_data = lead.model_dump(exclude_unset=True)
    from datetime import datetime, timezone
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    client = get_client_for_user(user.token)
    if "status" in update_data:
        old_lead = get_lead_status(client, lead_id)
        old_status = old_lead.data["status"]
        new_status = update_data["status"]

        if old_status != new_status:
            create_activity(client, {
                "lead_id": lead_id,
                "user_id": user.id,
                "type": "status_change",
                "content": f"Status changed from {old_status} to {new_status}",
                "metadata": {"from": old_status, "to": new_status},
            })

    response = db_update_lead(client, lead_id, update_data)
    return response.data

@app.delete("/leads/{lead_id}")
@limiter.limit("60/minute")
def delete_lead(request: Request, lead_id: str, user=Depends(require_admin)):
    client = get_client_for_user(user.token)
    response = db_delete_lead(client, lead_id)
    return {"deleted": True, "id": lead_id}

@app.get("/me")
@limiter.limit("60/minute")
def get_me(request: Request, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    profile = get_profile(client, user.id)
    return profile.data

@app.get("/leads/{lead_id}/activities")
@limiter.limit("60/minute")
def get_lead_activities(request: Request, lead_id: str, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    response = db_get_lead_activities(client, lead_id)
    return response.data

@app.post("/leads/{lead_id}/notes")
@limiter.limit("60/minute")
def create_note(request: Request, lead_id: str, note: NoteCreate, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    create_activity(client, {
        "lead_id": lead_id,
        "user_id": user.id,
        "type": "note",
        "content": note.content,
    })
    return {"status": "ok"}

@app.get("/leads/export")
@limiter.limit("5/minute")
def export_leads(request: Request, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    response = get_all_leads(client)
    leads = response.data

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=REQUIRED_COLUMNS)
    writer.writeheader()
    for lead in leads:
        writer.writerow({k: lead.get(k, "") for k in writer.fieldnames})

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads_export.csv"},
    )

REQUIRED_COLUMNS = [
    "first_name", "last_name", "title", "org", "email", "phone", "phone_2",
    "linkedin", "location", "industry", "function", "status", "next_action", "due_date",
    "revenue", "currency"
]
VALID_STATUSES = {"New", "Contacted", "Follow-up", "Won", "Lost"}

@app.get("/leads/import-template")
@limiter.limit("5/minute")
def import_template(request: Request, user=Depends(get_current_user)):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(REQUIRED_COLUMNS)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads_import_template.csv"},
    )

LINKEDIN_COLUMN_MAP = {
    "First Name": "first_name",
    "Last Name": "last_name",
    "Title": "title",
    "Company": "org",
    "Email": "email",
    "Phone Number 1": "phone",
    "Phone Number 2": "phone_2",
    "Phone Number": "phone",
    "Profile Url": "linkedin",
}
LINKEDIN_REQUIRED_HEADERS = {"First Name", "Last Name", "Company"}


def normalize_linkedin_row(raw_row: dict) -> dict:
    mapped = {col: None for col in REQUIRED_COLUMNS}
    for linkedin_col, internal_col in LINKEDIN_COLUMN_MAP.items():
        val = raw_row.get(linkedin_col)
        if val is not None and val != "":
            mapped[internal_col] = val
    mapped["status"] = "New"  # LinkedIn exports don't have a status column
    return mapped


def cell_to_str(value):
    return "" if value is None else str(value).strip()


@app.post("/leads/import")
def import_leads(file: UploadFile = File(...), user=Depends(get_current_user)):
    if file.filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(file.file.read()))
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    else:
        content = file.file.read().decode("utf-8")
        all_rows = list(csv.reader(io.StringIO(content)))

    if not all_rows:
        raise HTTPException(status_code=400, detail="File is empty")

    raw_headers = list(all_rows[0])
    data_rows = all_rows[1:]
    header_set = set(h for h in raw_headers if h is not None)

    if raw_headers == REQUIRED_COLUMNS:
        reader = [dict(zip(REQUIRED_COLUMNS, row)) for row in data_rows]
    elif LINKEDIN_REQUIRED_HEADERS.issubset(header_set):
        raw_dicts = [dict(zip(raw_headers, row)) for row in data_rows]
        reader = [normalize_linkedin_row(r) for r in raw_dicts]
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unrecognized file format. Headers must match either the app template {REQUIRED_COLUMNS} "
                   f"or a LinkedIn export (needs at least: {sorted(LINKEDIN_REQUIRED_HEADERS)}). Got: {raw_headers}"
        )

    client = get_client_for_user(user.token)
    def process_stream():
        imported = []
        errors = []
        total = len(reader)
        
        if total == 0:
            yield json.dumps({"type": "complete", "imported_count": 0, "errors": []}) + "\n"
            return

        for i, row in enumerate(reader, start=2):
            first_name = cell_to_str(row.get("first_name"))
            status = cell_to_str(row.get("status"))

            if not first_name:
                errors.append(f"Row {i}: 'first_name' is required, row rejected")
                yield json.dumps({"type": "progress", "processed": i - 1, "total": total, "percentage": int((i - 1) / total * 100)}) + "\n"
                continue

            if status not in VALID_STATUSES:
                errors.append(f"Row {i}: '{status}' is not a valid status {sorted(VALID_STATUSES)}, row rejected")
                yield json.dumps({"type": "progress", "processed": i - 1, "total": total, "percentage": int((i - 1) / total * 100)}) + "\n"
                continue

            lead_data = {col: (cell_to_str(row.get(col)) or None) for col in REQUIRED_COLUMNS}
            lead_data["first_name"] = first_name
            lead_data["status"] = status

            result = db_create_lead(client, lead_data)
            new_lead = result.data[0]

            create_activity(client, {
                "lead_id": new_lead["id"],
                "user_id": user.id,
                "type": "created",
                "content": "Lead created via import",
            })

            imported.append(new_lead["id"])
            
            yield json.dumps({"type": "progress", "processed": i - 1, "total": total, "percentage": int((i - 1) / total * 100)}) + "\n"

        yield json.dumps({"type": "complete", "imported_count": len(imported), "errors": errors}) + "\n"

    return StreamingResponse(process_stream(), media_type="application/x-ndjson")

@app.post("/leads/{lead_id}/attachments")
def upload_attachment(lead_id: str, file: UploadFile = File(...), user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    file_bytes = file.file.read()
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB.")
    safe_filename = sanitize_filename(file.filename)
    storage_path = f"{lead_id}/{safe_filename}"

    upload_file_to_storage(storage_path, file_bytes, file.content_type)

    result = create_attachment_record(client, {
        "lead_id": lead_id,
        "file_name": file.filename,
        "storage_path": storage_path,
        "uploaded_by": user.id,
    })

    create_activity(client, {
        "lead_id": lead_id,
        "user_id": user.id,
        "type": "note",
        "content": f"Attached file: {file.filename}",
    })

    return result.data


@app.get("/leads/{lead_id}/attachments")
def list_attachments(lead_id: str, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    response = get_lead_attachments(client, lead_id)
    return response.data


@app.get("/attachments/{attachment_id}/download")
def download_attachment(attachment_id: str, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    record = get_attachment(client, attachment_id)
    storage_path = record.data["storage_path"]

    signed_url = get_signed_attachment_url(storage_path)

    return {"url": signed_url["signedURL"]}


@app.delete("/attachments/{attachment_id}")
def delete_attachment(attachment_id: str, user=Depends(get_current_user)):
    client = get_client_for_user(user.token)
    record = get_attachment(client, attachment_id)
    storage_path = record.data["storage_path"]

    delete_file_from_storage(storage_path)
    delete_attachment_record(client, attachment_id)

    return {"deleted": True}

@app.get("/leads/import-template-xlsx")
@limiter.limit("5/minute")
def import_template_xlsx(request: Request, user=Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    ws.append(REQUIRED_COLUMNS)

    for col_idx, header in enumerate(REQUIRED_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    status_col_index = REQUIRED_COLUMNS.index("status") + 1
    status_col_letter = ws.cell(row=1, column=status_col_index).column_letter

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(sorted(VALID_STATUSES))}"',
        allow_blank=False,
    )
    dv.error = "Please select a valid status from the dropdown."
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add(f"{status_col_letter}2:{status_col_letter}1000")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leads_import_template.xlsx"},
    )